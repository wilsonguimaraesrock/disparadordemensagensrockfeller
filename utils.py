"""
WhatsApp API Sender - Utilities Module
Funções auxiliares para o sistema de envio em massa
"""

import os
import re
import json
import random
import time
import logging
from typing import List, Dict, Any, Optional
from pathlib import Path

import openpyxl
from openpyxl import load_workbook


def setup_logging(level: int = logging.DEBUG) -> None:
    """
    Configura o sistema de logging da aplicação
    
    Args:
        level: Nível de logging (padrão: INFO)
    """
    log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    
    # Configurar logging para arquivo
    log_filename = f"app_{time.strftime('%Y%m%d')}.log"
    
    logging.basicConfig(
        level=level,
        format=log_format,
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler()  # Também exibir no console
        ]
    )
    
    # Reduzir verbosidade de bibliotecas externas
    logging.getLogger('urllib3').setLevel(logging.WARNING)
    logging.getLogger('requests').setLevel(logging.WARNING)


def load_config(config_file: str = 'config.json') -> Dict[str, Any]:
    """
    Carrega as configurações da API do arquivo JSON
    
    Args:
        config_file: Caminho para o arquivo de configuração
        
    Returns:
        Dict com as configurações
        
    Raises:
        FileNotFoundError: Se o arquivo não for encontrado
        json.JSONDecodeError: Se o JSON for inválido
    """
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"Arquivo de configuração não encontrado: {config_file}")
    
    try:
        with open(config_file, 'r', encoding='utf-8') as file:
            config = json.load(file)
        
        # Validar campos obrigatórios
        required_fields = ['provider', 'base_url']
        missing_fields = [field for field in required_fields if field not in config]
        
        if missing_fields:
            raise ValueError(f"Campos obrigatórios ausentes no config.json: {', '.join(missing_fields)}")
        
        # Carregar variáveis de ambiente se especificado
        if 'env_token' in config:
            env_var = config['env_token']
            token_from_env = os.getenv(env_var)
            if token_from_env:
                config['token'] = token_from_env
        
        return config
        
    except json.JSONDecodeError as e:
        raise json.JSONDecodeError(f"Erro ao decodificar JSON do arquivo {config_file}: {e}")


def load_contacts_from_excel(file_path: str) -> List[Dict[str, str]]:
    """
    Carrega os contatos de uma planilha Excel
    
    Colunas esperadas:
    - Nome (opcional)
    - Numero (obrigatório)
    
    Args:
        file_path: Caminho para o arquivo Excel
        
    Returns:
        Lista de dicionários com os contatos
        
    Raises:
        FileNotFoundError: Se o arquivo não for encontrado
        Exception: Para outros erros de leitura
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Arquivo Excel não encontrado: {file_path}")
    
    contacts = []
    debug_logs = []
    logger = logging.getLogger(__name__)
    
    try:
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        
        # Encontrar cabeçalhos
        headers = {}
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        logger.info(f"Cabeçalho bruto lido da planilha: {first_row}")

        for idx, header in enumerate(first_row):
            if header:
                header_lower = str(header).lower().strip()
                logger.info(f"Processando cabeçalho: '{header}' -> '{header_lower}'")
                if 'nome' in header_lower:
                    headers['nome'] = idx
                elif 'numero' in header_lower or 'telefone' in header_lower or 'phone' in header_lower:
                    headers['numero'] = idx
        
        logger.info(f"Mapeamento de cabeçalhos final: {headers}")

        # Verificar se campos obrigatórios foram encontrados
        if 'numero' not in headers:
            logger.error("A coluna 'Numero' é obrigatória e não foi encontrada nos cabeçalhos.")
            raise ValueError("Coluna 'Numero' não encontrada na planilha. Verifique se o nome da coluna está correto.")
        
        logger.info(f"Cabeçalhos encontrados com sucesso: {headers}")
        
        # Ler dados
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            logger.debug(f"[DEBUG] Lendo linha {row_idx}: {row}")

            # Pular linhas completamente vazias
            if not any(row):
                logger.debug(f"Linha {row_idx} ignorada: completamente vazia.")
                continue

            try:
                # Extrair dados da linha
                raw_numero = row[headers['numero']]
                logger.debug(f"[DEBUG] Linha {row_idx} - Valor bruto de 'numero': '{raw_numero}' (Tipo: {type(raw_numero)})")

                numero = str(raw_numero).strip() if raw_numero is not None else ''
                logger.debug(f"[DEBUG] Linha {row_idx} - Valor de 'numero' após strip: '{numero}'")
                
                # Pular linhas onde o número é efetivamente vazio
                if not numero:
                    logger.warning(f"Linha {row_idx} ignorada: 'numero' está vazio ou contém apenas espaços.")
                    continue
                
                contact = {
                    'numero': numero
                }
                
                # Campo opcional: nome
                if 'nome' in headers and row[headers.get('nome')] is not None:
                    contact['nome'] = str(row[headers['nome']]).strip()
                
                contacts.append(contact)
                logger.info(f"[SUCCESS] Contato adicionado da linha {row_idx}: {contact}")
                
            except Exception as e:
                logger.error(f"Erro ao processar linha {row_idx}: {row}. Erro: {e}", exc_info=True)
                continue
        
        workbook.close()
        logger.info(f"Carregados {len(contacts)} contatos da planilha")
        return contacts
        
    except Exception as e:
        logger.error(f"Erro ao carregar planilha: {e}")
        raise


def validate_phone_number(phone: str) -> Optional[str]:
    """
    Valida e formata um número de telefone para o padrão E.164 (+55DDDNumero).
    
    Tenta corrigir formatos comuns, como:
    - (21) 99999-8888
    - 21 99999 8888
    - 5521999998888
    - 21999998888
    
    Args:
        phone: Número de telefone para validar e formatar.
    
    Returns:
        O número formatado em E.164 se for válido ou corrigível, senão None.
    """
    if not phone:
        return None
    
    # 1. Limpeza profunda: remover tudo que não for dígito
    cleaned_phone = re.sub(r'\D', '', str(phone))
    
    # 2. Casos comuns de correção
    # Se o número começar com '55' e tiver 12 ou 13 dígitos (55 + DDD + 8/9 dígitos), está ok
    if cleaned_phone.startswith('55') and 12 <= len(cleaned_phone) <= 13:
        return f"+{cleaned_phone}"
    
    # Se tiver 10 ou 11 dígitos (DDD + número), adicionar '55'
    if 10 <= len(cleaned_phone) <= 11:
        return f"+55{cleaned_phone}"
    
    # 3. Se o número começar com '+' e já estiver no formato correto
    if phone.startswith('+') and 12 <= len(cleaned_phone) <= 13:
        return f"+{cleaned_phone}"
    
    # 4. Se nenhuma regra corresponder, o número é inválido
    return None


def normalize_phone_number(phone: str) -> str:
    """
    Normaliza um número de telefone para o formato padrão
    
    Args:
        phone: Número original
        
    Returns:
        str: Número normalizado
    """
    # Remover caracteres especiais
    clean_phone = re.sub(r'[^\d+]', '', phone)
    
    # Se não começar com +, adicionar +55 se necessário
    if not clean_phone.startswith('+'):
        if not clean_phone.startswith('55'):
            clean_phone = '55' + clean_phone
        clean_phone = '+' + clean_phone
    
    return clean_phone


def wait_random_interval(min_seconds: int = 5, max_seconds: int = 20) -> int:
    """
    Aguarda um intervalo aleatório entre envios
    
    Args:
        min_seconds: Tempo mínimo em segundos
        max_seconds: Tempo máximo em segundos
        
    Returns:
        int: Tempo aguardado em segundos
    """
    wait_time = random.randint(min_seconds, max_seconds)
    time.sleep(wait_time)
    return wait_time


def format_file_size(size_bytes: int) -> str:
    """
    Formata o tamanho de um arquivo em formato legível
    
    Args:
        size_bytes: Tamanho em bytes
        
    Returns:
        str: Tamanho formatado (ex: "2.5 MB")
    """
    if size_bytes == 0:
        return "0 B"
    
    size_names = ["B", "KB", "MB", "GB"]
    i = 0
    size = float(size_bytes)
    
    while size >= 1024.0 and i < len(size_names) - 1:
        size /= 1024.0
        i += 1
    
    return f"{size:.1f} {size_names[i]}"


def validate_media_file(file_path: str) -> Dict[str, Any]:
    """
    Valida um arquivo de mídia
    
    Args:
        file_path: Caminho para o arquivo
        
    Returns:
        Dict com informações do arquivo ou erro
    """
    result = {
        'valid': False,
        'error': None,
        'size': 0,
        'type': None,
        'extension': None
    }
    
    try:
        if not os.path.exists(file_path):
            result['error'] = 'Arquivo não encontrado'
            return result
        
        # Verificar tamanho
        size = os.path.getsize(file_path)
        result['size'] = size
        
        # Verificar se não está vazio
        if size == 0:
            result['error'] = 'Arquivo vazio'
            return result
        
        # Verificar extensão
        extension = Path(file_path).suffix.lower()
        result['extension'] = extension
        
        # Tipos de arquivo suportados
        supported_extensions = {
            # Imagens
            '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp',
            # Vídeos
            '.mp4', '.avi', '.mov', '.wmv', '.flv', '.webm',
            # Áudios
            '.mp3', '.wav', '.ogg', '.m4a', '.aac',
            # Documentos
            '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt'
        }
        
        if extension not in supported_extensions:
            result['error'] = f'Tipo de arquivo não suportado: {extension}'
            return result
        
        # Determinar tipo
        if extension in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
            result['type'] = 'image'
        elif extension in ['.mp4', '.avi', '.mov', '.wmv', '.flv', '.webm']:
            result['type'] = 'video'
        elif extension in ['.mp3', '.wav', '.ogg', '.m4a', '.aac']:
            result['type'] = 'audio'
        else:
            result['type'] = 'document'
        
        # Verificar tamanho máximo (50MB)
        max_size = 50 * 1024 * 1024  # 50MB
        if size > max_size:
            result['error'] = f'Arquivo muito grande: {format_file_size(size)} (máximo: 50MB)'
            return result
        
        result['valid'] = True
        return result
        
    except Exception as e:
        result['error'] = f'Erro ao validar arquivo: {str(e)}'
        return result


def create_sample_excel(filename: str = 'sample.xlsx') -> None:
    """
    Cria um arquivo Excel de exemplo com a estrutura esperada
    
    Args:
        filename: Nome do arquivo a ser criado
    """
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Contatos"
        
        # Cabeçalhos
        headers = ['Nome', 'Numero']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Dados de exemplo
        sample_data = [
            ['João Silva', '+5521999998888'],
            ['Maria Santos', '+5511888887777'],
            ['Pedro Costa', '21777776666'],
            ['Ana Lima', '11666665555']
        ]
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        print(f"✅ Arquivo de exemplo criado: {filename}")
        
    except Exception as e:
        print(f"❌ Erro ao criar arquivo de exemplo: {e}")


def display_progress(current: int, total: int, message: str = '') -> None:
    """
    Exibe uma barra de progresso simples no terminal
    
    Args:
        current: Número atual
        total: Total de itens
        message: Mensagem adicional
    """
    if total == 0:
        return
    
    percentage = (current / total) * 100
    bar_length = 30
    filled_length = int(bar_length * current // total)
    
    bar = '█' * filled_length + '-' * (bar_length - filled_length)
    
    print(f'\r[{bar}] {percentage:.1f}% ({current}/{total}) {message}', end='', flush=True)
    
    if current == total:
        print()  # Nova linha ao final